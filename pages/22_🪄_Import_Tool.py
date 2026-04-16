import streamlit as st
from google import genai
from google.genai import types
from pypdf import PdfReader
import asyncio
import os
import io
import time
from datetime import datetime
import json
import pandas as pd
import re
import nest_asyncio
import pandas as pd
from openpyxl import load_workbook
nest_asyncio.apply()

st.set_page_config(layout="wide")
st.title("✨ Import Tool :blue-badge[:material/experiment: Beta]")

# Configuration des APIs
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")

# Initialisation des clients async
gemini_client = genai.Client(api_key=GOOGLE_API_KEY)

# Prompt personnalisé
custom_prompt = """
Vous êtes un agent d’extraction documentaire spécialisé dans les plans d’actions de transition écologique des collectivités, y compris les PCAET.

Contexte du fichier en entrée
Le texte fourni est un document de plan d’actions d’une collectivité. Un plan est structuré en axes, sous axes, actions et parfois sous actions. Le contenu peut être issu d’un PDF converti avec des artefacts de mise en page. Certaines rubriques comme budget, service pilote ou statut ne sont pas toujours explicites.

Objectif
Analyser le texte ci dessous et extraire toutes les actions, en reconstruisant la hiérarchie axe puis sous axe puis action, et en ajoutant des sous actions si nécessaire. Axes, sous axes et actions sont obligatoires. Les sous actions sont optionnels.

Sortie attendue
1 Répondre uniquement avec un tableau JSON valide
2 Ne rien ajouter avant ni après le JSON
3 Ne pas utiliser de balises Markdown

Schéma des objets du tableau
Chaque entrée du tableau est un objet avec exactement ces champs
[
 "axe",
 "sous-axe",
 "titre",
 "description",
 "sous-actions",
 "direction ou service pilote",
 "personne pilote",
 "budget",
 "statut"
]

Types et formats attendus
• "axe" est une chaîne
• "sous-axe" est une chaîne
• "titre" est une chaîne courte (300 caractères maximum). Si le titre issu du texte source dépasse 300 caractères, conserver uniquement les 300 premiers caractères significatifs dans "titre" et reporter le reste dans "description"
• "description" est une chaîne
• "sous-actions" est une liste de chaînes. Si aucune sous action ne s’impose, mettre une liste vide []
• "direction ou service pilote" est une chaîne
• "personne pilote" est une chaîne
• "budget" est soit la valeur vide "", soit un entier sans séparateur d’espace
• "statut" est une chaîne

Définitions opérationnelles
• Plan. Ensemble structuré d’orientations et de mesures d’une collectivité
• Axe. Grande orientation stratégique du plan. Exemple "Vers une mobilité vertueuse et réfléchie"
• Sous axe. Déclinaison thématique d’un axe. Exemple "Mettre en œuvre les conditions favorables à des déplacements plus sobres"
• Action. Mesure opérationnelle unique qui peut être mise en œuvre et suivie. Elle a un titre court et, uniquement si nécessaire, une description synthétique apportant des informations complémentaires au titre
• Sous action. Etape ou brique concrète qui détaille la mise en œuvre d’une action. Les sous actions sont listées dans "sous-actions"

Hiérarchie et numérotation
1 Conserver strictement les libellés exacts du texte source lorsque la numérotation et les titres existent
2 Lorsque le texte ne fournit pas de numérotation explicite, construire une numérotation stable et cohérente selon la règle suivante
   On note les axes "n".
   On note les sous axes "n.X".
   On note les actions "n.X.Y"
3 "axe" doit être formaté exactement "Axe n : Titre de l’axe"
4 "sous-axe" doit être formaté exactement "n.X  Titre du sous-axe"
6 "titre" doit être formaté "n.X.Y Titre de l’action"
7 Un sous axe doit avoir un nom complet. Il ne peut pas être uniquement un nombre
8 Pour un même identifiant hiérarchique le libellé doit être identique partout

Tâches obligatoires et ordre d’exécution
1 Normalisation du texte source
   • Retirer uniquement les artefacts manifestes de conversion comme "Unnamed" ou des mots isolés insérés au milieu d’une phrase
   • Conserver l’orthographe et les majuscules des noms propres et sigles
2 Relevé de structure
   • Repérer les axes puis les sous axes
3 Extraction des actions
   • Lister chaque action avec un titre court et une description synthétique fidèle au texte
   • Lorsque le texte présente des puces, des sous parties ou des verbes d’exécution multiples rattachés à une même action, créer des sous actions dans "sous-actions" comme une liste de chaînes
4 Rattachement hiérarchique
   • Associer chaque action à son sous axe et à son axe
5 Complétude des champs
   • Remplir "direction ou service pilote", "personne pilote", "budget" et "statut" uniquement si l’information est explicite et non ambiguë
   • Lorsque le texte présente pour une action des dates ou un calendrier (sans libellé de statut explicite pour cette action), vous pouvez déduire « statut » uniquement parmi « À venir » et « En cours » en vous appuyant sur ces dates et la date du jour ({date_du_jour}). Si seule une année est mentionnée sans mois précis (par exemple "2026"), considérer le statut comme « En cours » si cette année correspond à l'année actuelle, et « À venir » si elle est dans le futur
6 Validation du format
   • Produire un JSON valide
   • Vérifier que chaque objet contient exactement les champs définis
   • Si une information manque, la laisser à "" sauf "sous-actions" qui doit être une liste vide et "budget" qui doit être "" ou un entier
7 Dé duplication
   • Si deux entrées décrivent la même action, conserver une seule entrée avec la description la plus complète
8 Couverture
   • Parcourir tout le texte fourni et extraire l’ensemble des actions identifiables

Règles générales
1 Ne jamais inventer des informations ou des chiffres
2 Ne pas réécrire le sens de la "description". La nettoyer uniquement pour supprimer des artefacts évidents
3 "statut" ne peut prendre que l’une des valeurs suivantes sinon ""
   ["À venir", "À discuter", "En cours", "Réalisé", "En retard", "En pause", "Bloqué"]
4 "direction ou service pilote" contient uniquement des organismes ou services. "personne pilote" contient uniquement des noms de personnes
5 Majuscules. Mettre une majuscule au premier mot de chaque champ texte. Conserver les majuscules des noms propres et des sigles. Supprimer les espaces superflus au début et à la fin
6 Respect strict des libellés existants pour axes et sous axes lorsque fournis. En l’absence de libellé explicite, créer un libellé concis et fidèle au contenu
7 Ordre de tri. Le tableau doit être trié selon la hiérarchie axe puis sous axe puis ordre des actions
8 "titre" et "description" ne doivent jamais contenir les mêmes informations. La description apporte un complément au titre. Si le titre suffit à décrire l’action et qu’il n’y a rien de pertinent à ajouter, laisser "description" à ""

Exemples de bonne structure de plan
Exemple de titres hiérarchiques attendus quand le texte les fournit
Axe 1 : Une transition construite de manière transversale
1.1 S’appuyer sur un pilotage et des coopérations stables
1.1.1 Définir un portage politique fort
1.2 Impliquer tous les publics dans les transitions
Axe 2 : Vers un territoire rural affirmé aux multiples atouts en faveur du climat
2.1 Soutenir une agriculture paysanne
Axe 3 : Vers des équipements de qualité thermique et écologique
3.1 Concevoir des bâtiments publics de qualité une normalité
Axe 4 : Vers une mobilité vertueuse et réfléchie
4.2 Mettre en œuvre les conditions favorables à des déplacements plus sobres

Exemple de bonne extraction avec sous actions
Texte source
"Réduire l’autosolisme. Développer la pratique du covoiturage en s’appuyant tout d’abord sur des services existants mais aussi en mettant en place des infrastructures permettant de diversifier les offres
• S’appuyer sur l’offre existante proposée par Blablacar Daily pour le covoiturage domicile travail
• Déployer des lignes de covoiturage à haut niveau de service et les aménagements associés
• Réfléchir à des solutions d’autopartage en boucle"

Extraction attendue pour une action située dans le sous axe "4.2 Mettre en œuvre les conditions favorables à des déplacements plus sobres"
{
 "axe": "Axe 4  Vers une mobilité vertueuse et réfléchie",
 "sous-axe": "4.2  Mettre en œuvre les conditions favorables à des déplacements plus sobres",
 "titre": "4.2.1 Réduire l’autosolisme",
 "description": "Développer la pratique du covoiturage en s’appuyant sur des services existants et en mettant en place des infrastructures qui diversifient l’offre",
 "sous-actions": [
   "S’appuyer sur l’offre existante proposée par Blablacar Daily pour le covoiturage domicile travail",
   "Déployer des lignes de covoiturage à haut niveau de service et les aménagements associés",
   "Réfléchir à des solutions d’autopartage en boucle"
 ],
 "direction ou service pilote": "",
 "personne pilote": "",
 "budget": "",
 "statut": ""
}

Précisions sur le nettoyage minimal
• Retirer les mentions "Unnamed"
• Corriger les espaces multiples
• Conserver la ponctuation et les capitales des noms propres et sigles
• Ne pas corriger l’orthographe sauf artefacts de conversion manifestes

Consignes de saisie de champs
1 "direction ou service pilote" et "personne pilote" doivent contenir uniquement le nom de l’entité ou de la personne sans préposition. Exemple "SNCF" et non "Avec la SNCF"
2 **En cas de pluralité d’entités pour "direction ou service pilote" et/ou "personne pilote", les lister séparées par une virgule et un espace**
3 "budget" ne doit contenir que des chiffres sans séparateur ou la valeur vide
4 Si "statut" n’est pas exactement dans la liste autorisée, laisser ""

Rappel de robustesse
• Si le document fournit des numérotations et des titres, les réutiliser strictement
• Si des titres existent sans numéro, générer des numéros cohérents et stables
• Si la position d’une action parmi plusieurs sous axes demeure ambiguë, laisser vides les champs d’appartenance incertains plutôt que de forcer un rattachement

Jusqu’à présent, le prompt décrivait les règles générales d’extraction. Si le champ suivant n’est pas vide, vous devez impérativement tenir compte des précisions spécifiques ci-dessous.  
Elles peuvent modifier ou affiner l’interprétation de la structure du plan. Elles prévalent sur les règles générales lorsqu’il existe une contradiction ou une ambiguïté.

--- Précisions spécifiques (à appliquer strictement si présentes) qui prennent le dessus sur les règles générales ---
{precisions}
--- Fin des précisions spécifiques ---


Voici le texte à analyser :
{texte_pdf_a_analyser}  
"""

# Prompt de vérification 1 : vérifie la qualité de l'extraction
prompt_verif_1 = """
Tu es un agent de validation d’extractions documentaires extrêmement strict.

Contexte
Le document source accessible plus bas est un plan d’action.
Une première IA a déjà extrait une série d’actions. 

Structure attendue des actions
Les actions sont repérables car elles commencent par un identifiant numérique du type :
1.1.1 Titre de l’action; Description; Sous actions; Direction ou service pilote; Statut; Budget; Personne pilote; etc.
Tout ce qui suit une action appartient à cette action jusqu’au prochain identifiant du même type ou la fin du texte.

Objectif
1 Tu dois identifier chaque action dans le texte fourni
2 Pour chaque action, parcourir le document source via le file search.
3 Retrouver le passage correspondant à cette action dans le plan d’action.
4 Vérifier la fidélité de l’extraction pour cette action.
5 Attribuer un score de confiance entre 0 et 100 pour chaque identifiant d’action.

Critères de jugement
Tu dois juger uniquement sur
• omissions de texte significatives
• reformulations textuelles (changement de vocabulaire ou de structure de phrase)
Les ajouts de vocabulaire non présents dans le texte source sont considérés comme des reformulations.

Règles de notation
Tu dois attribuer pour chaque action un score entier entre 0 et 100, noté score, qui reflète la fidélité au texte source.

Guides de notation
• 100  texte quasi identique au texte source, aucune information manquante ni reformulation significative
• 90 à 99  quelques reformulations légères, aucun changement de sens, pas d’omission d’information importante
• 70 à 89  plusieurs reformulations ou petites omissions, mais le sens global reste correct
• 30 à 69  omissions importantes et ou nombreuses reformulations qui altèrent le texte
• 1 à 29  action très éloignée du contenu du document source
• 0  action hors sujet ou ne correspondant pas au document source

Changement de sens
• Si tu détectes un changement de sens, même partiel, le score doit chuter fortement en dessous de 70.
• Si le sens est largement incorrect ou trompeur, le score doit être inférieur ou égal à 30.

Contraintes pour le score
• Le score doit être un entier compris entre 0 et 100.
• Si le calcul te conduirait en dehors de ces bornes, ramène systématiquement le score dans l’intervalle.
• La notation doit suivre l’esprit du barème ci dessus et être stricte.

Format de sortie
Tu dois répondre uniquement avec un objet JSON strict représentant un dictionnaire :
• Les clés sont les identifiants des actions, sous forme d'un int qui se trouve entre les | au début du titre (exemple : 12)
• Les valeurs sont des objets contenant :
    - un entier entre 0 et 100 représentant le score de confiance
    - une explication très courte (quelques mots maximum) uniquement si le score est strictement inférieur à 90
    - si le score est supérieur ou égal à 90, l’explication doit être une chaîne vide ""

Exemples d’explications acceptables :
"omissions partielles"
"reformulation légère"
"altération mineure du sens"
"omissions + reformulation"

Exemple de format attendu :
{{
  "1": {{ "score": 95, "explication": "" }},
  "2": {{ "score": 82, "explication": "omissions partielles" }}
}}

Contraintes supplémentaires
• Ne pas recopier de longs extraits du document source.
• Ne pas citer le texte du plan.
• Ne pas ajouter de commentaires, d’explications ou de texte en dehors du JSON.
• Ne pas ajouter d’autres clés que les identifiants des actions.

Voici le texte extrait par l'IA : 
{reponse_ia}

Voici le texte original : 
{texte_pdf_a_analyser}
"""

# Prompt d'amélioration : améliore les actions à faible score
prompt_upgrade_1 = """
Vous êtes un agent d’extraction documentaire spécialisé dans les plans d’actions de transition écologique des collectivités, y compris les PCAET.

Contexte
On vous fournit :
1) Une liste d’actions ciblées que l’on souhaite extraire ou corriger.
2) Le texte source complet du plan d’actions (issu d’un PDF parfois bruité).

Vous NE devez travailler QUE sur les actions explicitement listées ci dessous.

Actions ciblées à traiter
Ces actions sont données sous forme de titres d'actions :

-------- DEBUT LISTE --------
{actions_a_ameliorer}
--------- FIN LISTE ---------

Texte source du plan d’actions. Il peut y avoir des artefacts de mise en page.

--------- TEXTE SOURCE ---------
{texte_pdf_a_analyser}
--------- FIN TEXTE SOURCE ---------

Objectif
Pour chaque action présente dans la liste "Actions ciblées à traiter" :
1) Parcourir le texte source.
2) Retrouver l’action correspondante à partir de son titre.
3) Extraire tous ses attributs en respectant strictement le schéma JSON décrit ci dessous.

Schéma des objets du tableau
La sortie doit être un tableau JSON. Chaque élément du tableau est un objet contenant exactement les champs suivants pour chaque index :

"titre"
"description"
"sous-actions"

Exemple :

{
  "12": {
    "titre": "1.4.1 Animer et suivre le COT et la démarche de transition écologique",
    "description": "Assurer le suivi des actions pilotées par les collègues ou d'autres acteurs, animer le Comité de pilotage, et assurer la mobilisation des élus.",
    "sous-actions": [
      "sous_action_1",
      "sous_action_2",
      "sous_action_3"
    ]
  }
}

L'index est donnée entre les | dans la liste en entrée


Types et formats attendus
• "titre" est une chaîne de la forme "n.X.Y Titre de l'action" qui doit correspondre à l’une des actions listées. Le titre ne doit pas dépasser 300 caractères. Si le titre issu du texte source dépasse 300 caractères, conserver uniquement les 300 premiers caractères significatifs dans "titre" et reporter le reste dans "description"
• "description" est une chaîne
• "sous-actions" est une liste de chaînes. Si aucune sous action ne s’impose, mettre []

Règles d’extraction spécifiques
1) Si l’information n’est pas explicitement présente dans le texte source, laisser ces champs à [] pour "sous-actions"
2) **SOYEZ COMPLETEMENT EXHAUSTIF SUR L'EXTRACTION NOTAMMENT DES DESCRIPTIONS ET SOUS-ACTIONS** 
3) Ne vous répétez pas entre les descriptions et les sous-actions, si certaines phrases s'apparentent à des sous-actions. Mettez les dans les sous-actions et non dans la description.
4) Le titre et la description ne doivent JAMAIS contenir les mêmes informations. La description ne doit apporter que des informations complémentaires au titre. Si le titre est suffisant, laisser "description" à "".

Nettoyage minimal
• Corriger les espaces multiples
• Retirer les artefacts manifestes ("Unnamed", numéros isolés sans sens, etc.)
• Ne pas réécrire le sens de la description, uniquement nettoyer les artefacts

Sortie attendue
1) Répondre uniquement avec un tableau JSON valide
2) Ne rien ajouter avant ni après le JSON
3) Ne pas utiliser de balises Markdown
4) Le tableau ne doit contenir QUE les actions demandées qui ont pu être retrouvées dans le texte source
"""


# Prompt de vérification qualitative finale
prompt_verif_quali = """
Vous êtes un auditeur qualité spécialisé dans les plans d’actions de transition écologique.

Contexte
On vous fournit une extraction déjà structurée avec les éléments suivants :
"axe", "sous-axe", "titre", "titre de la sous-action", "description", "direction ou service pilote", "personne pilote", "budget", "statut", "date de début", "date de fin" s'ils sont disponibles.
Les actions classiques ont "titre" rempli. Les sous-actions apparaissent comme des lignes séparées marquées [SA], avec "titre" vide et "titre de la sous-action" rempli.

Voici la sortie à évaluer
{reponse_ia}

Objectif
Votre travail n’est pas de corriger la sortie ni de la réécrire, mais de porter un jugement qualitatif sur sa qualité globale et de signaler les erreurs manifestes.

Axes d’évaluation
• Artefacts  vérifier qu’il ne subsiste pas d’artefacts évidents de conversion ou de mise en page comme "Unnamed", bouts de tableau, listes cassées, balises, répétitions absurdes, numérotations sans contenu.
• Cohérence sémantique  vérifier que chaque "description" a du sens, est compréhensible, et correspond à une action ou sous-action concrète de plan d’actions.
• Qualité des sous-actions  vérifier que les lignes marquées [SA] sont bien des sous-actions opérationnelles ou des étapes de mise en œuvre
• Cohérence hiérarchique  vérifier que "axe", "sous-axe" et "titre" sont cohérents entre eux, que la numérotation est plausible et stable, et que le contenu de l’action correspond bien à son axe et sous-axe.
• Champs pilotage, budget, statut  vérifier que "direction ou service pilote", "personne pilote", "budget" et "statut" ne semblent pas inventés, sont utilisés seulement lorsque l’information est explicitement plausible, et restent vides sinon.
• Doublons et éclatement inutile  vérifier qu’il n’y a pas de doublons évidents d’actions et que les actions ne sont pas artificiellement éclatées en plusieurs entrées identiques.
• Vérifier que les directions ou service pilote et personnes pilotes, si pluriel, sont des listes séparées par une virgule et un espace. S'il y a des tirets qui semble séparer deux entités **distinctes**, le relever.

Format de réponse attendu
• Répondre en français, sous forme de quelques lignes de texte libre.
• Commencer par un court avis global sur la qualité de l’extraction, par exemple "Extraction globalement cohérente avec quelques points à surveiller".
• Si tout est satisfaisant, le préciser explicitement, par exemple "Aucun problème majeur détecté".
• S’il existe des problèmes manifestes, les mentionner de manière ciblée en citant systématiquement la numérotation de l’action concernée, c’est à dire la partie "n.x.y" du champ "titre".
  Exemple  "1.2.4  description trop générale et peu opérationnelle" ou "4.1.3  présence probable d’artefacts de mise en page".
• Ne pas réécrire les actions et ne pas proposer de nouvelle version en JSON.
• Ne pas dépasser une dizaine de lignes.

Précision
• Les sous-actions apparaissent comme des lignes à part entière, avec "titre" vide et "titre de la sous-action" rempli.
• Les actions classiques ont "titre" rempli et "titre de la sous-action" vide.
"""

prompt_enrich_sous_actions = """
Vous êtes un agent d'enrichissement documentaire spécialisé dans les plans d'actions de transition écologique des collectivités, y compris les PCAET.

Contexte
On vous fournit :
1) Une liste numérotée de sous-actions, chacune rattachée à son action parente (indiquée entre crochets pour le contexte).
2) Le texte source complet du plan d'actions (issu d'un PDF parfois bruité).

Objectif
Pour chaque sous-action de la liste ci-dessous, parcourir le texte source et extraire les informations suivantes si elles sont explicitement présentes dans le document :
• "description" : une description synthétique et fidèle au texte source de la sous-action
• "personne_pilote" : le nom de la personne pilote ou référente de la sous-action.
• "statut" : le statut de la sous-action
• "date_debut" : la date de début
• "date_fin" : la date de fin ou échéance

Valeurs autorisées pour "statut"
["À venir", "À discuter", "En cours", "Réalisé", "En retard", "En pause", "Bloqué"]
Si le statut trouvé ne correspond à aucune de ces valeurs exactes, laisser "".

Règles strictes
1 Aucun champ n'est obligatoire. Si l'information n'est pas explicitement présente dans le texte source, laisser une chaîne vide "".
2 Ne jamais inventer d'informations, de dates ou de statuts.
3 Les dates doivent être au format JJ/MM/AAAA. Si seule l'année est disponible, utiliser 01/01/AAAA. Si seuls le mois et l'année sont disponibles, utiliser 01/MM/AAAA.
4 La description doit être fidèle au texte source. Ne pas réécrire le sens, uniquement nettoyer les artefacts de conversion.
5 Le titre de l'action parente entre crochets sert uniquement de contexte pour localiser la sous-action dans le document. Ne pas le reproduire dans la description.
6 Si une sous-action est trop générique ou introuvable dans le texte source, laisser tous les champs à "".
7 Une personne pilote doit forcément être une personne physique et NE PEUT PAS être une direction ou service
8 La description de la sous-action ne doit pas répéter le titre de la sous-action. Elle doit apporter des informations complémentaires. Si le titre est suffisant, laisser "description" à "".

Liste des sous-actions à enrichir

-------- DEBUT LISTE --------
{sous_actions_list}
--------- FIN LISTE ---------

Texte source du plan d'actions

--------- TEXTE SOURCE ---------
{texte_pdf_a_analyser}
--------- FIN TEXTE SOURCE ---------

Format de sortie
1 Répondre uniquement avec un objet JSON valide
2 Ne rien ajouter avant ni après le JSON
3 Ne pas utiliser de balises Markdown
4 Les clés sont les index numériques des sous-actions tels que fournis dans la liste

Exemple de format attendu
{{
  "0": {{"description": "Description trouvée dans le texte", "personne_pilote": "Jean Dupont", "statut": "En cours", "date_debut": "01/01/2025", "date_fin": "31/12/2025"}},
  "1": {{"description": "", "personne_pilote": "", "statut": "", "date_debut": "", "date_fin": ""}}
}}
"""

def extract_text_from_pdf(pdf_file):
    """Extrait le texte d'un fichier PDF"""
    try:
        pdf_reader = PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Erreur lors de l'extraction du PDF : {str(e)}"

def extract_text_from_csv(csv_file):
    try:
        df = pd.read_csv(csv_file, sep=';').fillna('')

        text = "# Fichier CSV\n\n"
        text += f"**Dimensions :** {len(df)} lignes × {len(df.columns)} colonnes\n\n"
        text += f"**Colonnes :** {', '.join(df.columns)}\n\n"
        text += "**Contenu complet :**\n\n"

        raw = df.to_string(index=False)
        raw = re.sub(r'\s+', ' ', raw)

        text += raw
        
        return text

    except Exception as e:
        return f"Erreur lors de la lecture du CSV : {str(e)}"

def extract_text_from_excel(excel_file):
    try:
        sheets = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')

        parts = []
        total_rows = 0
        for sheet_name, df in sheets.items():
            df = df.fillna('')
            total_rows += len(df)

            part = f"## Onglet : {sheet_name}\n\n"
            part += f"**Dimensions :** {len(df)} lignes × {len(df.columns)} colonnes\n\n"
            part += f"**Colonnes :** {', '.join(str(c) for c in df.columns)}\n\n"
            part += "**Contenu :**\n\n"

            raw = df.to_string(index=False)
            raw = re.sub(r'\s+', ' ', raw)
            part += raw

            parts.append(part)

        header = f"# Fichier Excel — {len(sheets)} onglet(s), {total_rows} lignes au total\n\n"
        return header + "\n\n---\n\n".join(parts)

    except Exception as e:
        return f"Erreur lors de la lecture du fichier Excel : {str(e)}"

def df_to_compact_text(df: pd.DataFrame, show_index: bool = True) -> str:
    """Convertit un dataframe en texte compact pour l'envoyer à Gemini.
    Gère deux formats : ancien (sous-actions en liste) et nouveau (titre de la sous-action en colonne)."""
    cols_expected = ["axe", "sous-axe"]
    for col in cols_expected:
        if col not in df.columns:
            raise ValueError(f"Colonne manquante dans le DataFrame : {col}")

    has_sous_actions_list = "sous-actions" in df.columns
    has_sous_action_col = "titre de la sous-action" in df.columns

    df_sorted = df.copy()
    df_sorted = df_sorted.sort_values(by=["axe", "sous-axe"]).reset_index(drop=True)

    parts = []

    def add_segment(segment: str):
        if not segment:
            return
        parts.append(segment.strip())

    for axe in df_sorted["axe"].dropna().unique():
        df_axe = df_sorted[df_sorted["axe"] == axe]
        add_segment(f"{axe} :")

        for sous_axe in df_axe["sous-axe"].dropna().unique():
            df_sous_axe = df_axe[df_axe["sous-axe"] == sous_axe]
            add_segment(f"Sous axe {sous_axe} :")

            for index_row, row in df_sous_axe.iterrows():
                champs_action = []

                titre = str(row.get("titre", "")).strip()
                titre_sa = str(row.get("titre de la sous-action", "")).strip() if has_sous_action_col else ""
                is_sous_action_row = has_sous_action_col and not titre and titre_sa

                if is_sous_action_row:
                    if show_index:
                        champs_action.append(f"| {index_row} | [SA] {titre_sa}")
                    else:
                        champs_action.append(f"[SA] {titre_sa}")
                elif titre:
                    if show_index:
                        champs_action.append(f"| {index_row} | {titre}")
                    else:
                        champs_action.append(f"{titre}")

                desc = str(row.get("description", "")).strip()
                if desc:
                    champs_action.append(f"{desc}")

                if has_sous_actions_list and not has_sous_action_col:
                    sous_actions = row.get("sous-actions", None)
                    if isinstance(sous_actions, (list, tuple)):
                        sa_clean = [str(sa).strip() for sa in sous_actions if str(sa).strip()]
                        if sa_clean:
                            champs_action.append("" + "; ".join(sa_clean))
                    elif isinstance(sous_actions, str) and sous_actions.strip():
                        champs_action.append(f"{sous_actions.strip()}")

                champs_optionnels = [
                    ("direction ou service pilote", "Direction ou service pilote"),
                    ("personne pilote", "Personne pilote"),
                    ("budget", "Budget"),
                    ("statut", "Statut"),
                    ("date de début", "Date de début"),
                    ("date de fin", "Date de fin"),
                ]

                for col, label in champs_optionnels:
                    if col in df_sous_axe.columns:
                        val = row.get(col, None)
                        if pd.notna(val):
                            val_str = str(val).strip()
                            if val_str:
                                champs_action.append(f"{label} {val_str}")

                if champs_action:
                    texte_action = "" + "; ".join(champs_action) + "."
                    add_segment(texte_action)

    return "\n".join(parts).strip()

def parse_json_response(result_text: str):
    """Parse une réponse JSON de Gemini en nettoyant les balises markdown"""
    cleaned_text = result_text.strip()
    if cleaned_text.startswith("```json"):
        cleaned_text = cleaned_text[7:]
    if cleaned_text.startswith("```"):
        cleaned_text = cleaned_text[3:]
    if cleaned_text.endswith("```"):
        cleaned_text = cleaned_text[:-3]
    cleaned_text = cleaned_text.strip()
    return json.loads(cleaned_text)


def normalize_sous_actions_value(val):
    """Retourne une liste de chaînes non vides pour la colonne « sous-actions ».

    Si le modèle renvoie une chaîne au lieu d'une liste, pandas/itération la parcourt
    caractère par caractère et déclenche à tort l'enrichissement des sous-actions.
    """
    if val is None:
        return []
    if isinstance(val, float) and pd.isna(val):
        return []
    if hasattr(val, "tolist") and not isinstance(val, (str, bytes, list, tuple, dict)):
        try:
            val = val.tolist()
        except Exception:
            return []
    if isinstance(val, (list, tuple)):
        return [str(x).strip() for x in val if str(x).strip()]
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ("[]", "null", "none", "nan"):
            return []
        if s.startswith("["):
            try:
                parsed = json.loads(s)
                if isinstance(parsed, list):
                    return [str(x).strip() for x in parsed if str(x).strip()]
            except json.JSONDecodeError:
                pass
        return []
    return []


def split_long_titles(df: pd.DataFrame, max_len: int = 300) -> pd.DataFrame:
    """Scinde les titres trop longs : conserve max_len chars dans le titre, reporte le reste dans description."""
    for col in ["titre", "titre de la sous-action"]:
        if col not in df.columns:
            continue
        for idx in df.index:
            val = df.at[idx, col]
            if not isinstance(val, str) or len(val) <= max_len:
                continue
            cut = val.rfind(" ", 0, max_len)
            if cut == -1:
                cut = max_len
            overflow = val[cut:].strip()
            df.at[idx, col] = val[:cut].strip()
            existing_desc = str(df.at[idx, "description"]).strip() if pd.notna(df.at[idx, "description"]) else ""
            if existing_desc:
                df.at[idx, "description"] = overflow + " - " + existing_desc
            else:
                df.at[idx, "description"] = overflow
    return df

def remplir_fichier_import(df: pd.DataFrame) -> io.BytesIO:
    """Remplit le fichier import avec les données du dataframe et retourne un BytesIO"""
    
    # 1 Charger le fichier source directement (sans copie sur disque)
    src = "utils/template_pa.xlsx"
    wb = load_workbook(src)
    ws = wb["Fichier dimport"]

    # 2 Mapping des colonnes Excel (lettre → nom de colonne df)
    mapping = {
        "A": "axe",
        "B": "sous-axe",
        "D": "titre",
        "E": "titre de la sous-action",
        "F": "description",
        "M": "direction ou service pilote",
        "N": "personne pilote",
        "X": "budget",
        "Y": "statut",
        "AA": "date de début",
        "AB": "date de fin",
    }

    # 3 Écrire les données à partir de la ligne 5
    start_row = 5

    for i, (_, row) in enumerate(df.iterrows(), start=start_row):
        for col_letter, df_col in mapping.items():
            value = row.get(df_col, "")
            ws[f"{col_letter}{i}"] = "" if pd.isna(value) else value

    # 4 Sauvegarder dans un BytesIO (en mémoire, pas sur disque)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def display_df_markdown(df: pd.DataFrame):
    """Affiche un dataframe en mode structuré (axes > sous-axes > actions et sous-actions)"""
    has_sous_action_col = "titre de la sous-action" in df.columns

    axes = df["axe"].unique()

    for axe in axes:
        st.markdown(f"### {axe}")

        df_axe = df[df["axe"] == axe]
        sous_axes = df_axe["sous-axe"].unique()

        for sous_axe in sous_axes:
            with st.expander(f"{sous_axe}", expanded=False):
                df_sous_axe = df_axe[df_axe["sous-axe"] == sous_axe]

                for _, row in df_sous_axe.iterrows():
                    titre = str(row.get("titre", "")).strip() if pd.notna(row.get("titre")) else ""
                    titre_sa = ""
                    if has_sous_action_col:
                        titre_sa = str(row.get("titre de la sous-action", "")).strip() if pd.notna(row.get("titre de la sous-action")) else ""

                    is_sous_action = has_sous_action_col and not titre and titre_sa

                    if is_sous_action:
                        sa_md = ""
                        sa_md += f"> **Sous-action :** {titre_sa}\n>\n"

                        if row.get("description") and str(row["description"]).strip():
                            sa_md += f"> **Description :** {row['description']}\n>\n"

                        if row.get("personne pilote") and str(row["personne pilote"]).strip():
                            sa_md += f"> **Personne pilote :** {row['personne pilote']}\n>\n"

                        if row.get("statut") and str(row["statut"]).strip():
                            sa_md += f"> **Statut :** {row['statut']}\n>\n"

                        if row.get("date de début") and str(row["date de début"]).strip():
                            sa_md += f"> **Date de début :** {row['date de début']}\n>\n"

                        if row.get("date de fin") and str(row["date de fin"]).strip():
                            sa_md += f"> **Date de fin :** {row['date de fin']}\n>\n"

                        st.markdown(sa_md)
                    else:
                        action_md = ""

                        if titre:
                            action_md += f"**Titre :** {titre}\n\n"

                        if row.get("description") and str(row["description"]).strip():
                            action_md += f"**Description :** {row['description']}\n\n"

                        if not has_sous_action_col:
                            sous_actions = row.get("sous-actions", [])
                            if sous_actions and len(sous_actions) > 0:
                                action_md += "**Sous-actions :**\n"
                                for sa in sous_actions:
                                    if sa and str(sa).strip():
                                        action_md += f"- {sa}\n"
                                action_md += "\n"

                        if row.get("direction ou service pilote") and str(row["direction ou service pilote"]).strip():
                            action_md += f"**Direction ou service pilote :** {row['direction ou service pilote']}\n\n"

                        if row.get("personne pilote") and str(row["personne pilote"]).strip():
                            action_md += f"**Personne pilote :** {row['personne pilote']}\n\n"

                        if row.get("budget") and str(row["budget"]).strip():
                            action_md += f"**Budget :** {row['budget']}\n\n"

                        if row.get("statut") and str(row["statut"]).strip():
                            action_md += f"**Statut :** {row['statut']}\n\n"

                        st.markdown(action_md)

                        if "score" in df.columns and row.get("score") is not None:
                            score = row["score"]
                            if row.get("amelioree", False):
                                st.info(f"FA consolidée. (confiance précédente: **{score}**)")
                            else:
                                explication = row.get("explication", "")
                                if explication and str(explication).strip():
                                    st.info(f"Confiance: **{score}** - {explication}")
                                else:
                                    st.info(f"Confiance: **{score}**")

                    st.markdown("---")

    st.markdown("#### ✅ Vue tableau")
    cols_to_drop = [c for c in ['score', 'explication', 'amelioree'] if c in df.columns]
    df_a_afficher = df.drop(columns=cols_to_drop).copy()
    st.dataframe(df_a_afficher, use_container_width=True, height=600)


async def query_gemini(user_prompt, model='gemini-3-pro-preview'):
    """Interroge Gemini avec streaming asynchrone"""
    start_time = time.time()
    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ✨ Gemini START ({model})")
    try:
        # Utiliser le streaming pour la réponse
        stream = await gemini_client.aio.models.generate_content_stream(
            model=model,
            contents=user_prompt,
            config=types.GenerateContentConfig(
                temperature=0.2,
                max_output_tokens=64000
            )
        )
        
        parts = []
        tokens = [0,0]
        last_chunk = None

        async for chunk in stream:
            if hasattr(chunk, 'text') and chunk.text:
                parts.append(chunk.text)
            last_chunk = chunk
        
        # Récupérer les tokens du dernier chunk
        if hasattr(last_chunk, 'usage_metadata') and hasattr(last_chunk.usage_metadata, 'candidates_token_count') and hasattr(last_chunk.usage_metadata, 'prompt_token_count'):
            tokens = [last_chunk.usage_metadata.candidates_token_count, last_chunk.usage_metadata.prompt_token_count]
        else:
            tokens = [0, 0]
        
        reponse = "".join(parts)
        elapsed = time.time() - start_time
        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ✅ Gemini END ({elapsed:.1f}s, {tokens} tokens)")
        return reponse, elapsed, tokens

    except Exception as e:
        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ❌ Gemini ERROR: {str(e)}")
        # Si le streaming ne fonctionne pas, fallback sur l'API standard
        try:
            response = await gemini_client.aio.models.generate_content(
                model=model,
                contents=user_prompt,
                config=types.GenerateContentConfig(
                    temperature=0.2,
                    max_output_tokens=64000
                )
            )
            elapsed = time.time() - start_time
            #Calcul des tokens
            if hasattr(response, 'usage_metadata') and hasattr(response.usage_metadata, 'candidates_token_count') and hasattr(response.usage_metadata, 'prompt_token_count'):
                tokens = [response.usage_metadata.candidates_token_count, response.usage_metadata.prompt_token_count]
            else:
                tokens = [0, 0]
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ✅ Gemini END (fallback) ({elapsed:.1f}s, {tokens[0]} tokens, {tokens[1]} tokens)")
            return str(response.text), elapsed, tokens
        
        except Exception as e2:
            elapsed = time.time() - start_time
            return f"Erreur Gemini: {str(e2)}", elapsed, [0, 0]


# ==========================
# Interface utilisateur
# ==========================

# Toggle pour le type de fichier
file_type = st.segmented_control(
    "Type de fichier à importer",
    options=["PDF", "CSV / Excel"],
    default="PDF"
)

if file_type == "PDF":
    uploaded_file = st.file_uploader(
        "Glissez-déposez votre fichier PDF ici",
        type=['pdf'],
        help="Sélectionnez un fichier PDF à analyser",
        key="pdf_uploader"
    )
else:
    uploaded_file = st.file_uploader(
        "Glissez-déposez votre fichier CSV ou Excel ici",
        type=['csv', 'xlsx', 'xls'],
        help="Fichiers acceptés : .csv, .xlsx, .xls. Les fichiers Excel multi-onglets sont supportés.",
        key="csv_excel_uploader"
    )

precisions = st.text_area(
    "Précisions",
    height=300,
    placeholder="Ajoutez des précisions supplémentaires si nécessaire. Vous pouvez ici définir une strucutre spécifique, certaines règles à respecter, donner du contexte, etc. Cliquez sur Ctrl+Enter pour valider."
)

# Choix du modèle Gemini
# gemini_model = st.segmented_control(
# "Modèle Gemini",
# options=["gemini-3-pro-preview", "gemini-2.5-pro"],
# default="gemini-2.5-pro"
# )
gemini_model = "gemini-2.5-pro"

avec_sous_actions = st.toggle("Avec sous-actions", value=True, help="Si désactivé, les sous-actions ne seront pas extraites ni enrichies.")
avec_verifications = st.toggle("Vérifications", value=True, help="Désactiver pour les fichiers simples (CSV, petits PDF). Ignore la vérification et la consolidation des fiches actions.")

# Mode test (tronque le texte à 10 000 caractères)
# mode_test = st.toggle("🧪 Mode test (30 000 caractères max)", value=False)
mode_test = False

total_tokens_consumed = [0, 0]

if uploaded_file is not None:
    st.success(f"✅ Fichier chargé : {uploaded_file.name}")
    
    start_button = st.button("🚀 Lancer l'analyse", type="primary")
    
    if start_button:
        # Extraction selon le type de fichier
        if file_type == "PDF":
            with st.spinner("📖 Extraction du texte du PDF..."):
                extracted_text = extract_text_from_pdf(uploaded_file)
        else:
            file_ext = uploaded_file.name.rsplit('.', 1)[-1].lower() if '.' in uploaded_file.name else ''
            if file_ext in ('xlsx', 'xls'):
                with st.spinner("🔍 Lecture du fichier Excel..."):
                    extracted_text = extract_text_from_excel(uploaded_file)
            else:
                with st.spinner("🔍 Lecture du fichier CSV..."):
                    extracted_text = extract_text_from_csv(uploaded_file)
        
        if extracted_text and not extracted_text.startswith("Erreur"):
            st.success(f"✅ Texte extrait : {len(extracted_text)} caractères")
            
            # Tronquer le texte en mode test
            if mode_test and len(extracted_text) > 30000:
                extracted_text = extracted_text[:30000]
                st.warning(f"🧪 Mode test activé : texte tronqué à 30 000 caractères")

            # ========================================
            # ÉTAPE 1 : Extraction initiale
            # ========================================
            st.markdown("---")
            st.markdown("## 🪄 Étape 1 : Définition de la structure et créations des fiches actions")
            
            user_prompt = custom_prompt.replace("{precisions}", precisions).replace("{texte_pdf_a_analyser}", extracted_text).replace("{date_du_jour}", datetime.now().strftime("%d/%m/%Y"))

            with st.spinner("🌀 Étape 1/5 : Définition de la structure et créations des fiches actions..."):
                gemini_result, elapsed_time, tokens_count = asyncio.run(query_gemini(user_prompt, gemini_model))
                total_tokens_consumed[0] += tokens_count[0]
                total_tokens_consumed[1] += tokens_count[1]
                st.info(f"✨ Extraction : {elapsed_time:.1f}s | Entrée : {tokens_count[1]:,} tokens | Sortie : {tokens_count[0]:,} tokens")
            
            if gemini_result and not gemini_result.startswith("Erreur"):
                try:
                    # Parser le JSON et créer le dataframe
                    data = parse_json_response(gemini_result)
                    df_actions = pd.DataFrame(data)
                    if not avec_sous_actions:
                        df_actions["sous-actions"] = [[] for _ in range(len(df_actions))]
                    elif "sous-actions" not in df_actions.columns:
                        df_actions["sous-actions"] = [[] for _ in range(len(df_actions))]
                    else:
                        df_actions["sous-actions"] = df_actions["sous-actions"].apply(
                            normalize_sous_actions_value
                        )
                    st.success(f"✅ {len(df_actions)} actions extraites")
                    st.dataframe(df_actions, use_container_width=True, height=400)
                    
                    if avec_verifications:
                        # ========================================
                        # ÉTAPE 2 : Vérification des scores
                        # ========================================
                        st.markdown("---")
                        st.markdown("## 🔍 Étape 2 : Vérification de la qualité des fiches actions")
                        
                        reponse_ia = df_to_compact_text(df_actions)
                        user_prompt_verif = prompt_verif_1.replace("{texte_pdf_a_analyser}", extracted_text).replace("{reponse_ia}", reponse_ia or "")
                        
                        with st.spinner("🌀 Étape 2/5 : Vérification de la qualité des fiches actions..."):
                            verif_result, elapsed_time, tokens_count = asyncio.run(query_gemini(user_prompt_verif, gemini_model))
                            total_tokens_consumed[0] += tokens_count[0]
                            total_tokens_consumed[1] += tokens_count[1]
                            st.info(f"✨ Vérification : {elapsed_time:.1f}s | Entrée : {tokens_count[1]:,} tokens | Sortie : {tokens_count[0]:,} tokens")
                    else:
                        st.markdown("---")
                        st.info("⏭️ Vérifications désactivées — étapes 2 et 3 ignorées.")
                        verif_result = "skip"

                    if verif_result and not verif_result.startswith("Erreur"):
                        try:
                            if avec_verifications:
                                # ========================================
                                # ÉTAPE 3 : Ajout des scores au dataframe
                                # ========================================
                                scores_data = parse_json_response(verif_result)
                                
                                df_actions["score"] = None
                                df_actions["explication"] = ""
                                df_actions["amelioree"] = False
                                
                                for idx_str, score_info in scores_data.items():
                                    idx = int(idx_str)
                                    if idx < len(df_actions):
                                        df_actions.at[idx, "score"] = score_info.get("score")
                                        df_actions.at[idx, "explication"] = score_info.get("explication", "")
                                
                                st.success(f"✅ Scores ajoutés pour {len(scores_data)} actions")
                                st.dataframe(df_actions[["titre", "score", "explication"]], use_container_width=True, height=300)
                                
                                # ========================================
                                # ÉTAPE 4 : Amélioration des actions à faible score
                                # ========================================
                                st.markdown("---")
                                st.markdown("## 🔧 Étape 3/5 : Consolidation des fiches actions")
                                
                                df_low_score = df_actions[df_actions["score"] < 90].copy()
                                
                                if len(df_low_score) > 0:
                                    st.warning(f"⚠️ {len(df_low_score)} actions avec un score < 90 à améliorer")
                                    
                                    BATCH_SIZE = 5
                                    low_score_indices = list(df_low_score.index)
                                    batches = [low_score_indices[i:i + BATCH_SIZE] for i in range(0, len(low_score_indices), BATCH_SIZE)]
                                    
                                    if len(batches) > 1:
                                        st.info(f"📦 Envoi de {len(batches)} batchs en parallèle à l'IA pour consolidation")
                                    else:
                                        st.info(f"📦 Envoi de {len(batches)} batch(s) en parallèle à l'IA pour consolidation")
                                    
                                    batch_prompts = []
                                    for batch_indices in batches:
                                        actions_a_ameliorer = ""
                                        for idx in batch_indices:
                                            row = df_actions.loc[idx]
                                            actions_a_ameliorer += f"|{idx}| {row['titre']}\n"
                                        
                                        batch_prompt = prompt_upgrade_1.replace("{texte_pdf_a_analyser}", extracted_text).replace("{actions_a_ameliorer}", actions_a_ameliorer)
                                        batch_prompts.append(batch_prompt)
                                    
                                    async def run_upgrade_batches():
                                        tasks = [query_gemini(prompt, gemini_model) for prompt in batch_prompts]
                                        return await asyncio.gather(*tasks, return_exceptions=True)
                                    
                                    with st.spinner(f"🌀 Étape 3/5 : Consolidation des fiches actions ({len(batches)} batches en parallèle)..."):
                                        batch_results = asyncio.run(run_upgrade_batches())
                                    
                                    total_upgraded = 0
                                    max_time = 0
                                    total_tokens = [0,0]
                                    all_errors = []
                                    
                                    for batch_idx, result in enumerate(batch_results):
                                        if isinstance(result, Exception):
                                            all_errors.append(f"Batch {batch_idx + 1}: {str(result)}")
                                            continue
                                        
                                        upgrade_result, elapsed_time, tokens_count = result
                                        max_time = max(max_time, elapsed_time)
                                        total_tokens[0] += tokens_count[0]
                                        total_tokens[1] += tokens_count[1]
                                        
                                        if upgrade_result and not upgrade_result.startswith("Erreur"):
                                            try:
                                                upgrade_data = parse_json_response(upgrade_result)
                                                
                                                for idx_str, item in upgrade_data.items():
                                                    idx = int(idx_str)
                                                    if idx < len(df_actions):
                                                        if "titre" in item:
                                                            df_actions.at[idx, "titre"] = item["titre"]
                                                        if "description" in item:
                                                            df_actions.at[idx, "description"] = item["description"]
                                                        if "sous-actions" in item and avec_sous_actions:
                                                            df_actions.at[idx, "sous-actions"] = normalize_sous_actions_value(item["sous-actions"])
                                                        df_actions.at[idx, "amelioree"] = True
                                                        total_upgraded += 1
                                            except Exception as e:
                                                all_errors.append(f"Batch {batch_idx + 1}: Parsing error - {str(e)}")
                                        else:
                                            all_errors.append(f"Batch {batch_idx + 1}: {upgrade_result}")
                                    
                                    total_tokens_consumed[0] += total_tokens[0]
                                    total_tokens_consumed[1] += total_tokens[1]
                                    st.info(f"✨ Consolidation : {max_time:.1f}s total | Entrée : {total_tokens[1]:,} tokens | Sortie : {total_tokens[0]:,} tokens")
                                    
                                    if total_upgraded > 0:
                                        st.success(f"✅ {total_upgraded} actions consolidées")
                                    
                                    if all_errors:
                                        for error in all_errors:
                                            st.error(f"❌ {error}")
                                else:
                                    st.success("✅ Toutes les actions ont un score > 90, pas de consolidation nécessaire")
                            else:
                                df_actions["score"] = None
                                df_actions["explication"] = ""
                                df_actions["amelioree"] = False
                            
                            # ========================================
                            # ÉTAPE 4 : Enrichissement des sous-actions
                            # ========================================
                            st.markdown("---")
                            st.markdown("## 🔎 Étape 4/5 : Enrichissement des sous-actions")

                            if not avec_sous_actions:
                                df_actions["sous-actions"] = [[] for _ in range(len(df_actions))]
                                st.info("⏭️ Sous-actions désactivées par l'utilisateur, étape ignorée.")

                            all_sous_actions = []
                            sa_parent_map = {}
                            global_idx = 0
                            for df_idx, row in df_actions.iterrows():
                                raw_val = row.get("sous-actions", None)
                                sous_actions = normalize_sous_actions_value(raw_val)
                                df_actions.at[df_idx, "sous-actions"] = sous_actions
                                for sa_pos, sa_str in enumerate(sous_actions):
                                    if sa_str:
                                        all_sous_actions.append({
                                            "global_idx": global_idx,
                                            "parent_idx": df_idx,
                                            "parent_titre": row["titre"],
                                            "sa_titre": sa_str,
                                            "sa_pos": sa_pos,
                                        })
                                        sa_parent_map[(df_idx, sa_pos)] = global_idx
                                        global_idx += 1

                            st.info(f"🔢 Sous-actions détectées : {len(all_sous_actions)}")

                            if len(all_sous_actions) > 0:
                                st.info(f"📋 {len(all_sous_actions)} sous-actions à enrichir")

                                SA_BATCH_SIZE = 30
                                sa_batches = [all_sous_actions[i:i + SA_BATCH_SIZE] for i in range(0, len(all_sous_actions), SA_BATCH_SIZE)]

                                st.info(f"📦 Envoi de {len(sa_batches)} batch(s) à l'IA pour enrichissement des sous-actions")

                                sa_batch_prompts = []
                                for batch in sa_batches:
                                    sous_actions_list = ""
                                    for sa_info in batch:
                                        sous_actions_list += f"{sa_info['global_idx']} | [Action parente : {sa_info['parent_titre']}] {sa_info['sa_titre']}\n"

                                    batch_prompt = prompt_enrich_sous_actions.replace("{sous_actions_list}", sous_actions_list).replace("{texte_pdf_a_analyser}", extracted_text)
                                    sa_batch_prompts.append(batch_prompt)

                                async def run_enrich_batches():
                                    tasks = [query_gemini(prompt, gemini_model) for prompt in sa_batch_prompts]
                                    return await asyncio.gather(*tasks, return_exceptions=True)

                                with st.spinner(f"🌀 Étape 4/5 : Enrichissement des sous-actions ({len(sa_batches)} batch(s) en parallèle)..."):
                                    sa_batch_results = asyncio.run(run_enrich_batches())

                                enrichment_data = {}
                                sa_max_time = 0
                                sa_total_tokens = [0, 0]
                                sa_errors = []

                                for batch_idx, result in enumerate(sa_batch_results):
                                    if isinstance(result, Exception):
                                        sa_errors.append(f"Batch {batch_idx + 1}: {str(result)}")
                                        continue

                                    enrich_result, elapsed_time, tokens_count = result
                                    sa_max_time = max(sa_max_time, elapsed_time)
                                    sa_total_tokens[0] += tokens_count[0]
                                    sa_total_tokens[1] += tokens_count[1]

                                    if enrich_result and not enrich_result.startswith("Erreur"):
                                        try:
                                            parsed = parse_json_response(enrich_result)
                                            for idx_str, fields in parsed.items():
                                                enrichment_data[int(idx_str)] = fields
                                        except Exception as e:
                                            sa_errors.append(f"Batch {batch_idx + 1}: Parsing error - {str(e)}")
                                    else:
                                        sa_errors.append(f"Batch {batch_idx + 1}: {enrich_result}")

                                total_tokens_consumed[0] += sa_total_tokens[0]
                                total_tokens_consumed[1] += sa_total_tokens[1]
                                st.info(f"✨ Enrichissement : {sa_max_time:.1f}s | Entrée : {sa_total_tokens[1]:,} tokens | Sortie : {sa_total_tokens[0]:,} tokens")

                                enriched_count = sum(1 for v in enrichment_data.values() if any(v.get(k) for k in ("description", "personne_pilote", "statut", "date_debut", "date_fin")))
                                if enriched_count > 0:
                                    st.success(f"✅ {enriched_count} sous-actions enrichies sur {len(all_sous_actions)}")

                                if sa_errors:
                                    for error in sa_errors:
                                        st.error(f"❌ {error}")

                                new_rows = []
                                for df_idx, row in df_actions.iterrows():
                                    action_row = row.to_dict()
                                    action_row["titre de la sous-action"] = ""
                                    action_row["date de début"] = ""
                                    action_row["date de fin"] = ""
                                    action_row.pop("sous-actions", None)
                                    new_rows.append(action_row)

                                    sous_actions_list_raw = row.get("sous-actions", [])
                                    if isinstance(sous_actions_list_raw, (list, tuple)):
                                        for sa_pos, sa in enumerate(sous_actions_list_raw):
                                            sa_str = str(sa).strip()
                                            if sa_str:
                                                gidx = sa_parent_map.get((df_idx, sa_pos))
                                                sa_enrichment = enrichment_data.get(gidx, {}) if gidx is not None else {}

                                                sa_row = {
                                                    "axe": row["axe"],
                                                    "sous-axe": row["sous-axe"],
                                                    "titre": row["titre"],
                                                    "titre de la sous-action": sa_str,
                                                    "description": sa_enrichment.get("description", ""),
                                                    "direction ou service pilote": "",
                                                    "personne pilote": sa_enrichment.get("personne_pilote", ""),
                                                    "budget": "",
                                                    "statut": sa_enrichment.get("statut", ""),
                                                    "date de début": sa_enrichment.get("date_debut", ""),
                                                    "date de fin": sa_enrichment.get("date_fin", ""),
                                                    "score": None,
                                                    "explication": "",
                                                    "amelioree": False,
                                                }
                                                new_rows.append(sa_row)

                                df_actions = pd.DataFrame(new_rows).reset_index(drop=True)
                                st.success(f"✅ Dataframe restructuré : {len(df_actions)} lignes (actions + sous-actions)")
                            else:
                                df_actions["titre de la sous-action"] = ""
                                df_actions["date de début"] = ""
                                df_actions["date de fin"] = ""
                                if "sous-actions" in df_actions.columns:
                                    df_actions = df_actions.drop(columns=["sous-actions"])
                                st.success("✅ Aucune sous-action à enrichir")

                            # ========================================
                            # ÉTAPE 5 : Vérification qualitative finale
                            # ========================================
                            st.markdown("---")
                            st.markdown("## ✅ Étape 5/5 : Vérifications finales")

                            df_actions = split_long_titles(df_actions)

                            for col in ["direction ou service pilote", "personne pilote"]:
                                if col in df_actions.columns:
                                    df_actions[col] = df_actions[col].apply(
                                        lambda x: x.replace("/", ", ") if isinstance(x, str) else x
                                    )

                            for col in df_actions.columns:
                                df_actions[col] = df_actions[col].apply(
                                    lambda x: re.sub(r' +', ' ', x).strip() if isinstance(x, str) else x
                                )

                            st.success("✅ Nettoyage effectué des colonnes pilotes")

                            reponse_ia_finale = df_to_compact_text(df_actions, show_index=False)
                            user_prompt_quali = prompt_verif_quali.replace("{reponse_ia}", reponse_ia_finale or "")

                            with st.spinner("🌀 Étape 5/5 : Analyse qualitative finale..."):
                                quali_result, elapsed_time, tokens_count = asyncio.run(query_gemini(user_prompt_quali, gemini_model))
                                total_tokens_consumed[0] += tokens_count[0]
                                total_tokens_consumed[1] += tokens_count[1]
                                st.info(f"✨ Vérifications finales : {elapsed_time:.1f}s | Entrée : {tokens_count[1]:,} tokens | Sortie : {tokens_count[0]:,} tokens")

                            # ========================================
                            # Affichage final
                            # ========================================
                            st.markdown("---")
                            st.markdown("## ✨ Plan final")
                            
                            # Afficher le résultat de la vérification qualitative
                            st.markdown(f"**Avis de l'IA** \n\n {quali_result}")

                            st.success(f"✅ Import réussi pour un cout d'environ {(10*total_tokens_consumed[0] + 2*total_tokens_consumed[1])/1000000:.2f} €")
                            
                            # Afficher le dataframe final en markdown
                            display_df_markdown(df_actions)

                            # Remplir le fichier import et proposer le téléchargement
                            try:
                                excel_data = remplir_fichier_import(df_actions)
                                
                                st.download_button(
                                    label="📥 Télécharger le fichier d'import rempli au format Excel",
                                    data=excel_data,
                                    file_name="import_plan_actions_" + pd.Timestamp.now().strftime('%Y%m%d_%H%M%S') + ".xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )
                            except Exception as e:
                                st.error(f"❌ Erreur lors du remplissage du fichier import : {str(e)}")
                            
                        except Exception as e:
                            st.error(f"❌ Erreur lors du parsing des scores : {str(e)}")
                            st.text(verif_result)
                    else:
                        st.error(f"❌ Erreur lors de la vérification : {verif_result}")
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors du parsing de l'extraction : {str(e)}")
                    st.text(gemini_result)
            else:
                st.error(f"❌ Erreur lors de l'extraction : {gemini_result}")
                
        
        else:
            st.error(f"❌ Erreur lors de l'extraction du texte du fichier")
            st.error(extracted_text)
else:
    label = "PDF" if file_type == "PDF" else "CSV ou Excel"
    st.info(f"👆 Veuillez charger un fichier {label} pour commencer")

